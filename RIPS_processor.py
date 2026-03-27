import os
import re
import time
import unicodedata
from pathlib import Path

from openpyxl import load_workbook, Workbook
from pdf2image import convert_from_path, pdfinfo_from_path
import pytesseract
from thefuzz import process, fuzz
from PIL import Image, ImageEnhance, ImageFilter

# ================== CONFIGURACIÓN ==================
BASE_DIR = Path(__file__).parent

INPUT_FOLDER = BASE_DIR / "input"
OUTPUT_FOLDER = BASE_DIR / "output"
DOCS_FOLDER = BASE_DIR / "docs"

EXCEL_DATOS = OUTPUT_FOLDER / "datos_rips.xlsx"
TABLA_REF = DOCS_FOLDER / "TablaReferencia_Municipio.xlsx"
EXCEL_PLANTILLA = DOCS_FOLDER / "plantilla.xlsm"
CACHE_FILE = DOCS_FOLDER / "genero_cache.xlsx"

FILA_FIJA = 1

# 👉 IMPORTANTE: ajustar según instalación local
pytesseract.pytesseract.tesseract_cmd = "tesseract"

# ================== FUNCIONES ==================

def quitar_tildes(txt):
    return ''.join(c for c in unicodedata.normalize('NFD', txt) if unicodedata.category(c) != 'Mn')

def normalizar_ciudad(txt):
    txt = quitar_tildes(txt).upper()
    txt = re.sub(r"[^A-Z0-9\s]", "", txt)
    return re.sub(r"\s+", " ", txt).strip()

def obtener_subtotal(texto):
    m = re.search(r"SUBTOTAL\s*[:\-]?\s*\$?\s*([\d.,]+)", texto)
    if not m:
        return None
    valor = m.group(1).replace(",", "").replace(".", "")
    return int(valor) if valor else None

def obtener_factura(pdf_path):
    return "CFY" + Path(pdf_path).stem

def mejorar_imagen(img):
    img = img.convert("L")
    img = ImageEnhance.Contrast(img).enhance(2.0)
    img = img.filter(ImageFilter.MedianFilter())
    return img

# ================== OCR ==================

def procesar_pdf(pdf):
    print(f"\n📄 Procesando: {pdf.name}")

    pages = convert_from_path(pdf, dpi=250)
    texto = ""

    for page in pages:
        page = mejorar_imagen(page)
        texto += pytesseract.image_to_string(page, lang="spa")

    texto = quitar_tildes(texto).upper()

    subtotal = obtener_subtotal(texto)
    factura = obtener_factura(pdf)

    return {
        "factura": factura,
        "subtotal": subtotal
    }

# ================== EXCEL ==================

def cargar_excel():
    if EXCEL_DATOS.exists():
        return load_workbook(EXCEL_DATOS)
    wb = Workbook()
    wb.active.title = "datos"
    return wb

def guardar_datos(datos):
    wb = cargar_excel()
    ws = wb.active

    fila = ws.max_row + 1 if ws.max_row > 1 else 2

    ws.cell(fila, 1, datos["factura"])
    ws.cell(fila, 2, datos["subtotal"])

    wb.save(EXCEL_DATOS)

# ================== MAIN ==================

def main():
    t0 = time.time()

    if not INPUT_FOLDER.exists():
        print("⚠️ Carpeta 'input' no encontrada")
        return

    pdfs = list(INPUT_FOLDER.glob("*.pdf"))

    if not pdfs:
        print("⚠️ No hay PDFs en la carpeta input")
        return

    print(f"📁 {len(pdfs)} archivos encontrados")

    for pdf in pdfs:
        datos = procesar_pdf(pdf)
        guardar_datos(datos)

    print(f"\n⏱️ Tiempo total: {time.time() - t0:.2f} segundos")

# ================== EJECUCIÓN ==================

if __name__ == "__main__":
    main()